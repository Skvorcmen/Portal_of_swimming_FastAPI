from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional, List
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.repositories.entry_repository import EntryRepository
from app.models import Entry
from app.core.exceptions import FileTooLargeError, InvalidFileError

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB


class EntryService:
    """Сервис для работы с заявками"""

    def __init__(self, session: AsyncSession):
        self.repo = EntryRepository(session)

    async def create_entry(
        self,
        competition_id: int,
        swim_event_id: int,
        athlete_id: int,
        entry_time: Optional[float] = None,
        status: str = "pending",
    ) -> Entry:
        """Создать новую заявку"""
        return await self.repo.create(
            competition_id=competition_id,
            swim_event_id=swim_event_id,
            athlete_id=athlete_id,
            entry_time=entry_time,
            status=status,
        )

    async def get_entry(self, entry_id: int) -> Optional[Entry]:
        """Получить заявку по ID"""
        return await self.repo.get_by_id(entry_id)

    async def get_entries_by_competition(self, competition_id: int) -> List[Entry]:
        """Получить все заявки для соревнования"""
        return await self.repo.get_by_competition(competition_id)

    async def get_entries_by_swim_event(self, swim_event_id: int) -> List[Entry]:
        """Получить заявки для дистанции"""
        return await self.repo.get_by_swim_event(swim_event_id)

    async def get_entries_by_athlete(self, athlete_id: int) -> List[Entry]:
        """Получить заявки спортсмена"""
        return await self.repo.get_by_athlete(athlete_id)

    async def update_entry_status(self, entry_id: int, status: str) -> Optional[Entry]:
        """Обновить статус заявки"""
        return await self.repo.update_status(entry_id, status)

    async def delete_entry(self, entry_id: int) -> bool:
        """Удалить заявку"""
        return await self.repo.delete(entry_id)

    async def import_from_excel(self, competition_id: int, file_content: bytes) -> dict:
        """Импорт заявок из Excel файла"""

        # Проверка размера файла
        if len(file_content) > MAX_FILE_SIZE:
            raise FileTooLargeError("File too large (max 5MB)")

        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            raise InvalidFileError(f"Error reading Excel file: {str(e)}")

        entries_data = []
        errors = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            try:
                swim_event_id = row[0]
                athlete_id = row[1]
                entry_time = row[2]
                status = row[3] if len(row) > 3 else "pending"

                if not swim_event_id or not athlete_id:
                    continue

                swim_event_id = int(swim_event_id)
                athlete_id = int(athlete_id)
                entry_time = float(entry_time) if entry_time else None

                if status not in ["pending", "confirmed", "rejected", "scratched"]:
                    status = "pending"

                entry = await self.create_entry(
                    competition_id=competition_id,
                    swim_event_id=swim_event_id,
                    athlete_id=athlete_id,
                    entry_time=entry_time,
                    status=status,
                )
                entries_data.append(entry)
            except (ValueError, TypeError) as e:
                errors.append(f"Row {row_idx}: Invalid data - {str(e)}")
                continue  # Продолжаем со следующей строкой
            except Exception as e:
                errors.append(f"Row {row_idx}: Database error - {str(e)}")
                continue  # Продолжаем со следующей строкой

        return {
            "message": f"Successfully created {len(entries_data)} entries",
            "entries": entries_data,
            "errors": errors if errors else None,
        }

    async def generate_excel_template(self, competition_id: int) -> BytesIO:
        """Генерация Excel шаблона для заявок"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"

        headers = ["swim_event_id", "athlete_id", "entry_time (сек)", "status"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=2, column=1, value="1")
        ws.cell(row=2, column=2, value="1")
        ws.cell(row=2, column=3, value="30.5")
        ws.cell(row=2, column=4, value="pending")

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
